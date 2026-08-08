import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

df = pd.read_csv("../data/retail_sales_clean.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet 1: Data (cleaned dataset that everything else references)
# ---------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Data"
cols = ["OrderID", "OrderDate", "Store", "Category", "SubCategory",
        "Sales", "Quantity", "Discount", "Profit", "Month", "Quarter", "Year", "MarginPct"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in df.iterrows():
    ws_data.append([row[c] for c in cols])
for i, c in enumerate(cols, 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 13
n_rows = len(df) + 1  # including header

# ---------------------------------------------------------------
# Sheet 2: Store_Summary  (SUMIFS formulas, not hardcoded)
# ---------------------------------------------------------------
ws_s = wb.create_sheet("Store_Summary")
stores = ["Dublin", "Cork", "Galway", "Limerick"]
years = [2024, 2025]

ws_s.append(["Store", "Year", "Revenue", "Profit", "MarginPct"])
for c in ws_s[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)

r = 2
for st in stores:
    for yr in years:
        ws_s.cell(row=r, column=1, value=st)
        ws_s.cell(row=r, column=2, value=yr)
        ws_s.cell(row=r, column=3,
            value=f'=ROUND(SUMIFS(Data!$F$2:$F${n_rows},Data!$C$2:$C${n_rows},A{r},Data!$L$2:$L${n_rows},B{r}),2)')
        ws_s.cell(row=r, column=4,
            value=f'=ROUND(SUMIFS(Data!$I$2:$I${n_rows},Data!$C$2:$C${n_rows},A{r},Data!$L$2:$L${n_rows},B{r}),2)')
        ws_s.cell(row=r, column=5, value=f'=IFERROR(ROUND(D{r}/C{r},4),0)')
        r += 1

# YoY growth block
ws_s.cell(row=1, column=7, value="Store").font = Font(bold=True, name="Arial")
ws_s.cell(row=1, column=8, value="2024 Revenue").font = Font(bold=True, name="Arial")
ws_s.cell(row=1, column=9, value="2025 Revenue").font = Font(bold=True, name="Arial")
ws_s.cell(row=1, column=10, value="YoY Growth %").font = Font(bold=True, name="Arial")
for c in range(7, 11):
    ws_s.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
    ws_s.cell(row=1, column=c).font = Font(bold=True, color=WHITE, name="Arial")

for i, st in enumerate(stores):
    row = 2 + i
    ws_s.cell(row=row, column=7, value=st)
    ws_s.cell(row=row, column=8,
        value=f'=SUMIFS($C$2:$C${r-1},$A$2:$A${r-1},G{row},$B$2:$B${r-1},2024)')
    ws_s.cell(row=row, column=9,
        value=f'=SUMIFS($C$2:$C${r-1},$A$2:$A${r-1},G{row},$B$2:$B${r-1},2025)')
    ws_s.cell(row=row, column=10, value=f'=ROUND((I{row}-H{row})/H{row}*100,2)')

for i in range(1, 11):
    ws_s.column_dimensions[get_column_letter(i)].width = 15

# Pivoted Store x Year revenue table (clean source for a clustered bar chart)
ws_s.cell(row=1, column=12, value="Store").font = Font(bold=True, color=WHITE, name="Arial")
ws_s.cell(row=1, column=13, value="2024").font = Font(bold=True, color=WHITE, name="Arial")
ws_s.cell(row=1, column=14, value="2025").font = Font(bold=True, color=WHITE, name="Arial")
for c in (12, 13, 14):
    ws_s.cell(row=1, column=c).fill = PatternFill("solid", fgColor=NAVY)
for i, st in enumerate(stores):
    row = 2 + i
    ws_s.cell(row=row, column=12, value=st)
    ws_s.cell(row=row, column=13, value=f'=H{row}')  # references the 2024 Revenue block above
    ws_s.cell(row=row, column=14, value=f'=I{row}')  # references the 2025 Revenue block above
for i in (12, 13, 14):
    ws_s.column_dimensions[get_column_letter(i)].width = 12

# ---------------------------------------------------------------
# Sheet 3: Category_Summary (top sub-categories by profit)
# ---------------------------------------------------------------
ws_c = wb.create_sheet("Category_Summary")
top10 = (df.groupby(["Category", "SubCategory"])
            .agg(Revenue=("Sales", "sum"), Profit=("Profit", "sum"))
            .reset_index()
            .sort_values("Profit", ascending=False)
            .head(10))

ws_c.append(["Category", "SubCategory", "Revenue", "Profit", "MarginPct"])
for c in ws_c[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)

row = 2
for _, tr in top10.iterrows():
    ws_c.cell(row=row, column=1, value=tr["Category"])
    ws_c.cell(row=row, column=2, value=tr["SubCategory"])
    ws_c.cell(row=row, column=3,
        value=f'=ROUND(SUMIFS(Data!$F$2:$F${n_rows},Data!$D$2:$D${n_rows},A{row},Data!$E$2:$E${n_rows},B{row}),2)')
    ws_c.cell(row=row, column=4,
        value=f'=ROUND(SUMIFS(Data!$I$2:$I${n_rows},Data!$D$2:$D${n_rows},A{row},Data!$E$2:$E${n_rows},B{row}),2)')
    ws_c.cell(row=row, column=5, value=f'=IFERROR(ROUND(D{row}/C{row},4),0)')
    row += 1
last_c_row = row - 1

# Cumulative % column for Pareto chart
ws_c.cell(row=1, column=6, value="CumulativePct").font = Font(bold=True, color=WHITE, name="Arial")
ws_c.cell(row=1, column=6).fill = PatternFill("solid", fgColor=NAVY)
for i in range(2, last_c_row + 1):
    ws_c.cell(row=i, column=6,
        value=f'=ROUND(SUM($D$2:D{i})/SUM($D$2:$D${last_c_row}),4)')

for i in range(1, 7):
    ws_c.column_dimensions[get_column_letter(i)].width = 18

# ---------------------------------------------------------------
# Sheet 4: Cork_vs_Others (headline insight)
# ---------------------------------------------------------------
ws_i = wb.create_sheet("Cork_Insight")
ws_i.append(["Store", "Electronics Revenue", "Store Total Revenue", "Electronics Share %"])
for c in ws_i[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, st in enumerate(stores):
    row = 2 + i
    ws_i.cell(row=row, column=1, value=st)
    ws_i.cell(row=row, column=2,
        value=f'=ROUND(SUMIFS(Data!$F$2:$F${n_rows},Data!$C$2:$C${n_rows},A{row},Data!$D$2:$D${n_rows},"Electronics"),2)')
    ws_i.cell(row=row, column=3,
        value=f'=ROUND(SUMIFS(Data!$F$2:$F${n_rows},Data!$C$2:$C${n_rows},A{row}),2)')
    ws_i.cell(row=row, column=4, value=f'=ROUND(B{row}/C{row}*100,2)')
for i in range(1, 5):
    ws_i.column_dimensions[get_column_letter(i)].width = 20

# ---------------------------------------------------------------
# Sheet 5: Dashboard (KPI cards + charts)
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)  # move to front
ws_d.sheet_view.showGridLines = False

ws_d.merge_cells("B2:K2")
ws_d["B2"] = "IRISH RETAIL SALES PERFORMANCE DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=20, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "Multi-Store Overview — Dublin | Cork | Galway | Limerick  (2024–2025)"
ws_d["B3"].font = Font(italic=True, size=12, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=11, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=22, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

kpi_card(ws_d, 2, "TOTAL REVENUE (2024-25)", "=ROUND(SUM(Data!F2:F%d),0)" % n_rows, '#,##0" €"')
kpi_card(ws_d, 4, "TOTAL PROFIT (2024-25)", "=ROUND(SUM(Data!I2:I%d),0)" % n_rows, '#,##0" €"')
kpi_card(ws_d, 6, "AVG MARGIN %", "=ROUND(SUM(Data!I2:I%d)/SUM(Data!F2:F%d),4)" % (n_rows, n_rows), '0.0%')
kpi_card(ws_d, 8, "DUBLIN YoY GROWTH", "=Store_Summary!J2", '0.0"%"')
kpi_card(ws_d, 10, "CORK ELECTRONICS SHARE", "=Cork_Insight!D3", '0.0"%"')

ws_d.row_dimensions[5].height = 18
ws_d.row_dimensions[6].height = 22
ws_d.row_dimensions[7].height = 22

# Store revenue bar chart (clustered: 2024 vs 2025 per store)
bar = BarChart()
bar.type = "col"
bar.grouping = "clustered"
bar.title = "Revenue by Store (2024 vs 2025)"
bar.style = 10
bar.y_axis.title = "Revenue (€)"
bar.x_axis.title = "Store"
data = Reference(ws_s, min_col=13, max_col=14, min_row=1, max_row=5)
cats = Reference(ws_s, min_col=12, min_row=2, max_row=5)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 9
ws_d.add_chart(bar, "B10")

# Pareto chart: profit bars + cumulative % line for top 10 sub-categories
pareto = BarChart()
pareto.type = "col"
pareto.title = "Top 10 Sub-Categories by Profit (Pareto)"
pareto.y_axis.title = "Profit (€)"
pdata = Reference(ws_c, min_col=4, min_row=1, max_row=last_c_row)
pcats = Reference(ws_c, min_col=2, min_row=2, max_row=last_c_row)
pareto.add_data(pdata, titles_from_data=True)
pareto.set_categories(pcats)
pareto.width, pareto.height = 16, 9

line = LineChart()
ldata = Reference(ws_c, min_col=6, min_row=1, max_row=last_c_row)
line.add_data(ldata, titles_from_data=True)
line.y_axis.axId = 200
line.y_axis.title = "Cumulative %"
line.y_axis.crosses = "max"
pareto += line
ws_d.add_chart(pareto, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 16
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Irish_Retail_Sales_Dashboard.xlsx")
print("Workbook saved.")
